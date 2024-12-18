from openpyxl import Workbook
from openpyxl.styles import Alignment
import numpy as np

aminoacidos_esenciales = ['histidina', 'isoleucina', 'leucina', 'lisina', 'metionina', 'fenilalanina', 'treonina', 'triptofano', 'valina']

def crear_tabla_calculos(titulo, nombres, fraccion_proteina, digestibilidad_proteina, contenido_aminoacidos, requerimientos, porcentajes_mezcla, aminoacidos_mezcla_gr_mezcla, fraccion_proteina_mezcla, puntaje_aminoacidos,  score_proteico, digestibilidad, PDCAAS):
    wb = Workbook()

    ws = wb.active

    N=3 # Espacio movil para encabezados

    # ENCABEZADO
    ws.cell(row=2, column=2, value=titulo)
    ws.cell(row=2, column=2).style = 'Headline 1'
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4+len(aminoacidos_esenciales))

    # TABLA 1 - Ingredientes
    ws.cell(row=2+N, column=2, value='INGRED.')
    ws.cell(row=2+N, column=2).style = 'Headline 3'
    ws.cell(row=2+N, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2+N, column=3, value='FRAC. PROT.')
    ws.cell(row=2+N, column=3).style = 'Headline 3'
    ws.cell(row=2+N, column=3).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2+N, column=4, value='DIGEST. PROT.')
    ws.cell(row=2+N, column=4).style = 'Headline 3'
    ws.cell(row=2+N, column=4).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1+N, column=5, value='CONTENIDO DE AMINOACIDOS. [mg por gr de proteína]')
    ws.cell(row=1+N, column=5).style = 'Headline 3'
    ws.cell(row=1+N, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1+N, start_column=5, end_row=1+N, end_column=4+len(aminoacidos_esenciales))

    for nc,aa_nombre in enumerate(aminoacidos_esenciales, start=5):
        ws.cell(row=2+N, column=nc, value=aa_nombre)


    for nf,id in enumerate(nombres.keys(), start=3):

        ws.cell(row=nf+N, column=2, value=nombres[id])

        ws.cell(row=nf+N, column=3, value=fraccion_proteina[id])

        ws.cell(row=nf+N, column=4, value=digestibilidad_proteina[id])

        for nc,aa in enumerate(contenido_aminoacidos[id], start=5):

            ws.cell(row=nf+N, column=nc, value=aa)


    # TABLA 2 - Requerimientos
    ws.cell(row=5+len(nombres)+N, column=5, value='REQUERIMIENTOS DE AMINOÁCIDOS ESCENCIALES [mg por gr proteína]')
    ws.cell(row=5+len(nombres)+N, column=5).style = 'Headline 3'
    ws.cell(row=5+len(nombres)+N, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=5+len(nombres)+N, start_column=5, end_row=5+len(nombres)+N, end_column=4+len(aminoacidos_esenciales))
    
    for nc,aa_nombre in enumerate(aminoacidos_esenciales, start=5):
        ws.cell(row=6+len(nombres)+N, column=nc, value=aa_nombre)

        ws.cell(row=7+len(nombres)+N, column=nc, value=requerimientos[0][nc-5])


    # SECCION RESULTADOS MEZCLA
    ws.cell(row=9+len(nombres)+N, column=2, value='RESULTADOS DE MEZCLA')
    ws.cell(row=9+len(nombres)+N, column=2).style = 'Headline 1'
    ws.cell(row=9+len(nombres)+N, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=9+len(nombres)+N, start_column=2, end_row=9+len(nombres)+N, end_column=4+len(aminoacidos_esenciales))

    # PROTEINA DE MEZCLA
    ws.cell(row=11+len(nombres)+N, column=3, value='FRAC. PROT. MEZCLA')
    ws.cell(row=11+len(nombres)+N, column=3).style = 'Headline 2'
    ws.cell(row=11+len(nombres)+N, column=3).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws.cell(row=12+len(nombres)+N, column=3, value=fraccion_proteina_mezcla)
    ws.cell(row=12+len(nombres)+N, column=3).number_format = '0.0%'


    # TABLA 3 - Porcentajes mezcla ingredientes
    ws.cell(row=2+N, column=7+len(aminoacidos_esenciales), value='FRACCION MEZCLA [%]')
    ws.cell(row=2+N, column=7+len(aminoacidos_esenciales)).style = 'Headline 2'
    ws.cell(row=2+N, column=7+len(aminoacidos_esenciales)).alignment = Alignment(horizontal="center", vertical="center")

    for nf,p in enumerate(porcentajes_mezcla,3):
        ws.cell(row=nf+N, column=7+len(aminoacidos_esenciales), value=p[0]*100)

    # TABLA 4 - Aminoácidos mezcla por gramo de mezcla
    ws.cell(row=10+len(nombres)+N, column=5, value='AMINOACIDOS EN MEZCLA  [mg por gr de mezcla]')
    ws.cell(row=10+len(nombres)+N, column=5).style = 'Headline 3'
    ws.cell(row=10+len(nombres)+N, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=10+len(nombres)+N, start_column=5, end_row=10+len(nombres)+N, end_column=4+len(aminoacidos_esenciales))
    
    for nc,aa_nombre in enumerate(aminoacidos_esenciales, start=5):
        ws.cell(row=11+len(nombres)+N, column=nc, value=aa_nombre)
        ws.cell(row=12+len(nombres)+N, column=nc, value=aminoacidos_mezcla_gr_mezcla[0][nc-5]  )
        ws.cell(row=12+len(nombres)+N, column=nc).number_format = '0.0000'
    

    # TABLA 5 - Aminoácidos mezcla por gr de proteína
    ws.cell(row=14+len(nombres)+N, column=5, value='AMINOACIDOS EN MEZCLA  [mg por gr de proteína en la mezcla]')
    ws.cell(row=14+len(nombres)+N, column=5).style = 'Headline 3'
    ws.cell(row=14+len(nombres)+N, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=14+len(nombres)+N, start_column=5, end_row=14+len(nombres)+N, end_column=4+len(aminoacidos_esenciales))
    
    for nc,aa_nombre in enumerate(aminoacidos_esenciales, start=5):
        ws.cell(row=15+len(nombres)+N, column=nc, value=aa_nombre)
        ws.cell(row=16+len(nombres)+N, column=nc, value=(aminoacidos_mezcla_gr_mezcla[0][nc-5])/fraccion_proteina_mezcla  )
        ws.cell(row=16+len(nombres)+N, column=nc).number_format = '0.0000'

    # TABLA 6 - AAS 
    ws.cell(row=18+len(nombres)+N, column=5, value='PUNTAJE DE AMINOACIDOS (AAS)')
    ws.cell(row=18+len(nombres)+N, column=5).style = 'Headline 3'
    ws.cell(row=18+len(nombres)+N, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=18+len(nombres)+N, start_column=5, end_row=18+len(nombres)+N, end_column=4+len(aminoacidos_esenciales))

    for nc,aa_nombre in enumerate(aminoacidos_esenciales, start=5):
        ws.cell(row=19+len(nombres)+N, column=nc, value=aa_nombre)
        ws.cell(row=20+len(nombres)+N, column=nc, value=puntaje_aminoacidos[0][nc-5] )
        ws.cell(row=20+len(nombres)+N, column=nc).number_format = '0.0000'

    # RESULTADO FINAL - DIGESTIBILIDAD Y PDCAAS
    ws.cell(row=23+len(nombres)+N, column=5, value='DIGESTIBILIDAD')
    ws.cell(row=23+len(nombres)+N, column=5).style = 'Headline 3'
    ws.cell(row=23+len(nombres)+N, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=24+len(nombres)+N, column=5, value=digestibilidad)
    ws.cell(row=24+len(nombres)+N, column=5).number_format = '0.0%'

    ws.cell(row=26+len(nombres)+N, column=5, value='SCORE PROTEICO')
    ws.cell(row=26+len(nombres)+N, column=5).style = 'Headline 3'
    ws.cell(row=26+len(nombres)+N, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=27+len(nombres)+N, column=5, value=score_proteico)
    ws.cell(row=27+len(nombres)+N, column=5).number_format = '0.0%'

    ws.cell(row=29+len(nombres)+N, column=5, value='PDCAAS')
    ws.cell(row=29+len(nombres)+N, column=5).style = 'Headline 3'
    ws.cell(row=29+len(nombres)+N, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=30+len(nombres)+N, column=5, value=PDCAAS)
    ws.cell(row=30+len(nombres)+N, column=5).number_format = '0.0%'

    return wb
